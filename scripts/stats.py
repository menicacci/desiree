import os
from scripts import utils, claim
from scripts.constants import Constants
import pandas as pd
from openpyxl import Workbook
from scripts.llm import llm_utils


def wrong_claims_prc(data_dir: str):
    model_answers_path = os.path.join(data_dir, Constants.Directories.ANSWERS)
    extracted_claims_path = os.path.join(data_dir, Constants.Filenames.CLAIMS)
    extracted_claims = claim.extract_claims_from_answers(model_answers_path, extracted_claims_path)

    claims_correctness_prc = {}
    ovr_c = 0
    ovr_w = 0

    for request_id, claims in extracted_claims.items():
        correct_claims = claims[Constants.Attributes.EXTRACTED_CLAIMS]
        wrong_claims = claims[Constants.Attributes.WRONG_CLAIMS]

        num_c_claims = len(correct_claims)
        num_w_claims = len(wrong_claims)

        claims_correctness_prc[request_id] = utils.divide_by_sum(num_c_claims, num_w_claims)

        ovr_c += num_c_claims
        ovr_w += num_w_claims

    return list(claims_correctness_prc.items()), utils.divide_by_sum(ovr_c, ovr_w)


def save(file_path: str, data_to_save: list, header: list):
    workbook = Workbook()
    sheet = workbook.active

    for idx, header_name in enumerate(header, start=1):
        sheet.cell(row=1, column=idx, value=header_name)

    for row_idx, data_dict in enumerate(data_to_save, start=2):
        for col_idx, key in enumerate(header, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=data_dict.get(key))

    workbook.save(file_path)


def write_ground_truth(gt_file_path: str, output_dir: str):
    utils.check_path(output_dir)

    data = pd.read_excel(gt_file_path, engine='odf', dtype={h: str for h in Constants.ColumnHeaders.TYPE_HEADER_STRUCTURE})
    for _, row in data.iterrows():
        request_id, output = [row[header_attr] for header_attr in Constants.ColumnHeaders.TYPE_HEADER_STRUCTURE]

        utils.write_file(str(output), os.path.join(output_dir, f"{request_id}.txt"))


def agglomerate_results(ground_truth_answers: dict, model_output: dict):
    return {
        request_id: (gt_answer, model_output[request_id])
        for request_id, gt_answer in ground_truth_answers.items()
        if request_id in model_output
    }


def compare_multiple_results(
        gt_path: str, 
        output_dirs: list[str],
        save_path: str,
        compare_function,
        stat_file_name=Constants.Filenames.STATS,
        opts=None
):
    gt_answers = llm_utils.read_model_output(gt_path)

    for output_dir in output_dirs:
        save_results_path = os.path.join(save_path, os.path.basename(output_dir))
        answer_path = os.path.join(output_dir, Constants.Directories.ANSWERS)
        stats_path = os.path.join(output_dir, stat_file_name)
        
        results = llm_utils.read_model_output(answer_path)
        agg_results = agglomerate_results(gt_answers, results)

        compare_function(agg_results, stats_path, save_results_path, opts)
